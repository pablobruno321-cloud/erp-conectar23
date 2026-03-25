"""
Módulo para crear backups automáticos en Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
from models import Cliente, Proveedor, Producto, Pedido, ItemPedido, Pago, Cobranza, ProveedorLogistico

def crear_backup_excel():
    """Crea un backup de todos los datos en Excel"""
    
    # Crear directorio si no existe
    backup_dir = Path('WEBAPPS/1-Conectar23')
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    # Nombre del archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    archivo = backup_dir / f'backup_{timestamp}.xlsx'
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía
    
    # Estilos
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== CLIENTES ====================
    ws = wb.create_sheet('Clientes')
    headers = ['ID', 'Nombre', 'CUIT', 'Teléfono', 'Email', 'Dirección', 'Saldo', 'Activo', 'Fecha Creación']
    ws.append(headers)
    
    for header in ws[1]:
        header.fill = header_fill
        header.font = header_font
        header.border = border
    
    for cliente in Cliente.query.all():
        ws.append([
            cliente.id,
            cliente.nombre,
            cliente.cuit,
            cliente.telefono,
            cliente.email,
            cliente.direccion,
            cliente.saldo,
            'Sí' if cliente.activo else 'No',
            cliente.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # ==================== PROVEEDORES ====================
    ws = wb.create_sheet('Proveedores')
    headers = ['ID', 'Nombre', 'CUIT', 'Teléfono', 'Email', 'Dirección', 'Saldo', 'Activo', 'Fecha Creación']
    ws.append(headers)
    
    for header in ws[1]:
        header.fill = header_fill
        header.font = header_font
        header.border = border
    
    for proveedor in Proveedor.query.all():
        ws.append([
            proveedor.id,
            proveedor.nombre,
            proveedor.cuit,
            proveedor.telefono,
            proveedor.email,
            proveedor.direccion,
            proveedor.saldo,
            'Sí' if proveedor.activo else 'No',
            proveedor.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # ==================== PRODUCTOS ====================
    ws = wb.create_sheet('Productos')
    headers = ['ID', 'Nombre', 'Proveedor', 'Clasificación', 'Variedad', 'Marca', 'Envase', 'Costo Unit.', 'Stock', 'Activo']
    ws.append(headers)
    
    for header in ws[1]:
        header.fill = header_fill
        header.font = header_font
        header.border = border
    
    for producto in Producto.query.all():
        ws.append([
            producto.id,
            producto.nombre,
            producto.proveedor.nombre if producto.proveedor else '-',
            producto.clasificacion,
            producto.variedad,
            producto.marca,
            producto.envase,
            producto.costo_unitario,
            producto.stock,
            'Sí' if producto.activo else 'No'
        ])
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # ==================== PEDIDOS ====================
    ws = wb.create_sheet('Pedidos')
    headers = ['ID', 'Número', 'Cliente', 'Fecha Venta', 'Costo Total', 'Precio Venta', 'Resultado', 'Estado', 'Mercado', 'Puesto']
    ws.append(headers)
    
    for header in ws[1]:
        header.fill = header_fill
        header.font = header_font
        header.border = border
    
    for pedido in Pedido.query.all():
        estado = 'Entregado' if pedido.entregado else ('Cargado' if pedido.cargado else 'Pendiente')
        ws.append([
            pedido.id,
            pedido.numero,
            pedido.cliente.nombre if pedido.cliente else '-',
            pedido.fecha_venta.strftime('%d/%m/%Y'),
            pedido.costo_total,
            pedido.precio_venta_total,
            pedido.resultado,
            estado,
            pedido.mercado,
            pedido.puesto
        ])
    
    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Guardar
    try:
        wb.save(archivo)
        return True, str(archivo)
    except Exception as e:
        return False, str(e)
